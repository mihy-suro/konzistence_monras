from typing import Optional, Tuple, List
import re
import openpyxl

def norm_text(s: str) -> str:
    """Normalizuje text hlavičky - odstraní speciální znaky, whitespace, převede na lowercase."""
    s = "" if s is None else str(s)
    # Odstranění newlines, tabulátorů a normalizace whitespace
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    # Odstranění jednotek v hranatých závorkách pro lepší matching
    s = re.sub(r"\s*\[[^\]]*\]", "", s).strip()
    # Nahrazení podtržítek mezerami pro konzistenci
    s = s.replace("_", " ")
    return s

def find_header_row_in_sheet(ws,
                             expected_norm: set,
                             max_rows: int,
                             min_hits: int,
                             min_ratio: float) -> Optional[int]:
    scan_rows = min(max_rows, ws.max_row or max_rows)
    for r in range(1, scan_rows + 1):
        row = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
        normed = [norm_text(v) for v in row]
        nonempty = [v for v in normed if v]
        if not nonempty:
            continue

        hits = sum(1 for v in normed if v in expected_norm)
        ratio = hits / max(1, len(nonempty))

        if hits >= min_hits and ratio >= min_ratio:
            return r
    return None

def detect_sheet_and_header(xlsx_path: str,
                            expected_header: List[str],
                            max_rows: int,
                            min_hits: int,
                            min_ratio: float) -> Tuple[str, int]:
    expected_norm = set(norm_text(x) for x in expected_header)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            hr = find_header_row_in_sheet(ws, expected_norm, max_rows, min_hits, min_ratio)
            if hr is not None:
                return name, hr
    finally:
        wb.close()

    raise RuntimeError(f"V souboru '{xlsx_path}' se nepodařilo najít list s očekávanou hlavičkou.")
